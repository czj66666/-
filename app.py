import streamlit as st
import pulp
import pandas as pd
import plotly.graph_objects as go
from io import BytesIO
import datetime
from openpyxl import Workbook

# ==================== 页面配置 ====================
st.set_page_config(page_title="Blueberry Pro v1.7", layout="wide")

st.markdown("""
<style>
[data-testid="stMetricValue"] { color: black !important; }
[data-testid="stMetricLabel"] { color: #1e3a8a !important; font-weight: bold; }
.stMetric { background-color: #f1f5f9; border-radius: 10px; border-left: 6px solid #1e3a8a; }
h1, h2, h3 { color: #1e3a8a; }
</style>
""", unsafe_allow_html=True)

# ==================== 肥料分组 ====================
MACRO_FERTILIZERS = [
    "Urea 尿素",
    "MAP 磷酸一铵",
    "MKP 磷酸二氢钾",
    "KNO3 硝酸钾",
    "K2SO4 硫酸钾",
    "Mg(NO3)2 硝酸镁",
    "MgSO4 硫酸镁",
    "Ca(NO3)2 硝酸钙",
    "AmSulphate 硫酸铵"
]

MICRO_FERTILIZERS = [
    "Iron 螯合铁",
    "MnSO4 硫酸锰",
    "ZnSO4 硫酸锌",
    "CuSO4·5H2O 硫酸铜",
    "Borax 硼砂",
    "Mo 钼酸铵"
]

# ==================== 初始化肥料库 ====================
if 'fert_lib' not in st.session_state:
    cols = ["NO3-N","NH4-N","P","K","Mg","Ca","Fe","SO4-S","Mn","Zn","Cu","B","Mo","Urea-N"]
    init_data = {
        "Urea 尿素": [0,0,0,0,0,0,0,0,0,0,0,0,0,0.46],
        "MAP 磷酸一铵": [0,0.12,0.266,0,0,0,0,0,0,0,0,0,0,0],
        "MKP 磷酸二氢钾": [0,0,0.227,0.299,0,0,0,0,0,0,0,0,0,0],
        "KNO3 硝酸钾": [0.135,0,0,0.38,0,0,0,0,0,0,0,0,0,0],
        "K2SO4 硫酸钾": [0,0,0,0.446,0,0,0,0.18,0,0,0,0,0,0],
        "Mg(NO3)2 硝酸镁": [0.10,0,0,0,0.09,0,0,0,0,0,0,0,0,0],
        "MgSO4 硫酸镁": [0,0,0,0,0.095,0,0,0.125,0,0,0,0,0,0],
        "Ca(NO3)2 硝酸钙": [0.11,0,0,0,0,0.16,0,0,0,0,0,0,0,0],
        "AmSulphate 硫酸铵": [0,0.21,0,0,0,0,0,0.24,0,0,0,0,0,0],
        "Iron 螯合铁": [0,0,0,0,0,0,0.13,0,0,0,0,0,0,0],
        "MnSO4 硫酸锰": [0,0,0,0,0,0,0,0.18,0.31,0,0,0,0,0],
        "ZnSO4 硫酸锌": [0,0,0,0,0,0,0,0.17,0,0.35,0,0,0,0],
        "Borax 硼砂": [0,0,0,0,0,0,0,0,0,0,0,0.11,0,0],
        "Mo 钼酸铵": [0,0,0,0,0,0,0,0,0,0,0,0,0.42,0],
        "CuSO4·5H2O 硫酸铜": [0,0,0,0,0,0,0,0.128,0,0,0.255,0,0,0],
    }
    st.session_state.fert_lib = pd.DataFrame.from_dict(init_data, orient='index', columns=cols).fillna(0.0)

# 统一名称，修复历史版本空格不一致
st.session_state.fert_lib.index = (
    st.session_state.fert_lib.index
    .str.replace(r"\s+", " ", regex=True)
    .str.strip()
)

# ==================== 基础计算函数 ====================
def calc_fertilizer_only(inputs, vol, rate):
    lib = st.session_state.fert_lib.fillna(0.0).to_dict('index')
    all_cols = st.session_state.fert_lib.columns.tolist()
    ppm = {col: 0.0 for col in all_cols}

    for name, kg in inputs.items():
        if name in lib and kg > 0:
            factor = (kg * 1_000_000 * rate) / vol
            for col in ppm.keys():
                ppm[col] += factor * float(lib[name][col])
    return ppm

def safe_calc(inputs, vol, rate, water, ec_factor):
    ppm = calc_fertilizer_only(inputs, vol, rate)
    all_cols = st.session_state.fert_lib.columns.tolist()

    res = {col: ppm[col] + water.get(col, 0.0) for col in all_cols if col != "Urea-N"}
    res["Urea-N"] = ppm["Urea-N"]

    meq = {
        "NH4+": res["NH4-N"] / 14.01,
        "K+": res["K"] / 39.1,
        "Ca2+": res["Ca"] / 20.04,
        "Mg2+": res["Mg"] / 12.15,
        "NO3-": res["NO3-N"] / 14.01,
        "H2PO4-": res["P"] / 30.97,
        "SO4 2-": res["SO4-S"] / 16.03,
        "HCO3-": water.get("HCO3", 0.0) / 61.02
    }

    s_cat = sum([meq["NH4+"], meq["K+"], meq["Ca2+"], meq["Mg2+"]])
    s_ani = sum([meq["NO3-"], meq["H2PO4-"], meq["SO4 2-"], meq["HCO3-"]])
    total_n = res["NO3-N"] + res["NH4-N"] + res["Urea-N"]
    est_ec = ((s_cat + s_ani) / 20) * ec_factor + water["EC"]

    return res, total_n, meq, est_ec, s_cat, s_ani

# ==================== 调酸统一计算 ====================
def get_water_for_calc(w_data, dosing_rate, tank_vol):
    acid_mode = st.session_state.get("acid_mode", "不调酸")
    target_pH = st.session_state.get("target_pH", 5.5)

    hco3_val = w_data.get("HCO3", 0.0)
    current_hco3_meq = hco3_val / 61.02

    if target_pH <= 5.0:
        target_residual_meq = 0.1
    elif target_pH <= 5.5:
        target_residual_meq = 0.1 + (target_pH - 5.0) * (0.4 / 0.5)
    elif target_pH <= 6.0:
        target_residual_meq = 0.5 + (target_pH - 5.5) * (0.5 / 0.5)
    else:
        target_residual_meq = 1.0

    acid_catalog = {
        "磷酸 (H3PO4)": {
            "mw": 98.0, "val": 1, "el": "P", "el_w": 30.97,
            "options": {
                "75%": {"density": 1.58, "purity": 0.75},
                "80%": {"density": 1.63, "purity": 0.80},
                "85%": {"density": 1.685, "purity": 0.85},
            }
        },
        "硫酸 (H2SO4)": {
            "mw": 98.07, "val": 2, "el": "SO4-S", "el_w": 32.06,
            "options": {
                "50%": {"density": 1.40, "purity": 0.50},
                "98%": {"density": 1.84, "purity": 0.98},
            }
        },
        "硝酸 (HNO3)": {
            "mw": 63.01, "val": 1, "el": "NO3-N", "el_w": 14.01,
            "options": {
                "30%": {"density": 1.18, "purity": 0.30},
                "40%": {"density": 1.25, "purity": 0.40},
                "55%": {"density": 1.33, "purity": 0.55},
                "68%": {"density": 1.41, "purity": 0.68},
            }
        }
    }

    base_water = dict(w_data)
    water_with_acid = dict(w_data)
    acid_list = st.session_state.get("acid_list", [])
    acid_detail_rows = []

    needed_meq_total = 0.0
    acid_L_per_bucket_total = 0.0
    acid_additions = {"P": 0.0, "SO4-S": 0.0, "NO3-N": 0.0}

    if acid_mode == "调酸":
        needed_meq_total = max(0.0, current_hco3_meq - target_residual_meq)
        active_acids = [a for a in acid_list if a.get("enabled", True) and a.get("share", 0) > 0]
        total_share = sum(a.get("share", 0) for a in active_acids)

        if needed_meq_total > 0 and total_share > 0:
            for item in active_acids:
                acid_name = item["acid_type"]
                conc_label = item.get("conc_label", "")
                share = item.get("share", 0) / total_share

                base = acid_catalog[acid_name]
                opt = base["options"][conc_label]
                density = opt["density"]
                purity = opt["purity"]

                acid_meq = needed_meq_total * share
                ml_per_m3 = (acid_meq * base["mw"]) / (density * purity * base["val"])
                nutrient_ppm = (acid_meq / base["val"]) * base["el_w"]
                acid_L_per_bucket = ml_per_m3 * tank_vol / (dosing_rate * 1_000_000)

                acid_additions[base["el"]] += nutrient_ppm
                acid_L_per_bucket_total += acid_L_per_bucket

                acid_detail_rows.append({
                    "酸种": acid_name,
                    "浓度": conc_label,
                    "分担比例(%)": round(share * 100, 1),
                    "中和碱度(meq/L)": round(acid_meq, 3),
                    "用量(ml/m³)": round(ml_per_m3, 2),
                    "单桶加酸(L)": round(acid_L_per_bucket, 3),
                    f"{base['el']}增加(ppm)": round(nutrient_ppm, 4)
                })

        water_with_acid["HCO3"] = target_residual_meq * 61.02
        for el, add_val in acid_additions.items():
            water_with_acid[el] = water_with_acid.get(el, 0.0) + add_val

    return (
        water_with_acid,
        base_water,
        acid_additions,
        current_hco3_meq,
        needed_meq_total,
        target_residual_meq,
        acid_L_per_bucket_total,
        acid_detail_rows
    )

# ==================== 显示与导出 ====================
def build_ppm_breakdown(res, inputs, vol, rate, base_water, acid_additions):
    fert_ppm = calc_fertilizer_only(inputs, vol, rate)
    rows = []
    for el in res.keys():
        base_val = 0.0 if el == "Urea-N" else float(base_water.get(el, 0.0))
        acid_val = 0.0 if el == "Urea-N" else float(acid_additions.get(el, 0.0))
        fert_val = float(fert_ppm.get(el, 0.0))
        total_val = float(res.get(el, 0.0))
        rows.append({
            "元素": el,
            "原水本底": round(base_val, 4),
            "酸带入": round(acid_val, 4),
            "肥料贡献": round(fert_val, 4),
            "总 ppm": round(total_val, 4)
        })
    return pd.DataFrame(rows)

def export_to_excel(solution_dict, acid_rows, res_df, meq, total_n, ec, sc, sa, raw_water=None):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Fertilizer Plan"
    ws1.append(["类别", "名称", "投料量", "单位"])
    for k, v in solution_dict.items():
        ws1.append(["肥料", k, round(v, 4), "kg"])
    if acid_rows:
        for row in acid_rows:
            ws1.append(["酸液", f"{row['酸种']} {row['浓度']}", round(row["单桶加酸(L)"], 4), "L"])

    ws2 = wb.create_sheet("Element PPM")
    ws2.append(["元素", "原水本底", "酸带入", "肥料贡献", "总 ppm"])
    for _, row in res_df.iterrows():
        ws2.append([
            row["元素"],
            round(row["原水本底"], 4),
            round(row["酸带入"], 4),
            round(row["肥料贡献"], 4),
            round(row["总 ppm"], 4)
        ])

    ws3 = wb.create_sheet("Ion Balance")
    ws3.append(["离子", "meq/L"])
    for k, v in meq.items():
        ws3.append([k, round(v, 4)])
    ws3.append([])
    ws3.append(["Σ 阳离子", round(sc, 4)])
    ws3.append(["Σ 阴离子", round(sa, 4)])
    ws3.append(["总氮", round(total_n, 4)])
    ws3.append(["预测EC", round(ec, 4)])

    ws4 = wb.create_sheet("Water Params")
    ws4.append(["参数", "数值"])
    if raw_water:
        for k, v in raw_water.items():
            ws4.append([k, v])

    return wb

def show_results(res, tn, meq, ec, sc, sa, final_dict, base_water=None, acid_additions=None, acid_rows=None, raw_water=None):
    if base_water is None:
        base_water = {}
    if acid_additions is None:
        acid_additions = {}
    if acid_rows is None:
        acid_rows = []

    df_res = build_ppm_breakdown(
        res=res,
        inputs=final_dict,
        vol=tank_vol,
        rate=dosing_rate,
        base_water=base_water,
        acid_additions=acid_additions
    )

    c1, c2, c3, c4, c5 = st.columns(5)
    c1.metric("总氮", f"{round(tn,1)} ppm")
    c2.metric("预测 EC", f"{round(ec,2)}")
    c3.metric("Σ 阳", round(sc,2))
    c4.metric("Σ 阴", round(sa,2))
    c5.metric("电荷差", round(sc-sa,2))

    st.divider()
    l, r = st.columns([1, 1.2])

    with l:
        st.subheader("元素 ppm 明细")
        st.dataframe(df_res, use_container_width=True, hide_index=True)

        st.subheader("投料方案")
        plan_rows = [{"类别": "肥料", "名称": k, "投料量": round(v, 4), "单位": "kg"} for k, v in final_dict.items()]
        if acid_rows:
            for row in acid_rows:
                plan_rows.append({
                    "类别": "酸液",
                    "名称": f"{row['酸种']} {row['浓度']}",
                    "投料量": round(row["单桶加酸(L)"], 4),
                    "单位": "L"
                })
        plan_df = pd.DataFrame(plan_rows)
        st.dataframe(plan_df, use_container_width=True, hide_index=True)

    with r:
        fig = go.Figure(data=[
            go.Bar(name='阳离子', x=['NH4+','K+','Ca2+','Mg2+'],
                   y=[meq['NH4+'], meq['K+'], meq['Ca2+'], meq['Mg2+']]),
            go.Bar(name='阴离子', x=['NO3-','H2PO4-','SO4 2-','HCO3-'],
                   y=[meq['NO3-'], meq['H2PO4-'], meq['SO4 2-'], meq['HCO3-']])
        ])
        fig.update_layout(height=350, barmode='group')
        st.plotly_chart(fig, use_container_width=True)

        if acid_rows:
            st.subheader("酸液明细")
            st.dataframe(pd.DataFrame(acid_rows), use_container_width=True, hide_index=True)

    wb = export_to_excel(final_dict, acid_rows, df_res, meq, tn, ec, sc, sa, raw_water=raw_water)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    st.download_button(
        "📥 下载完整Excel报告",
        buffer,
        file_name=f"Blueberry_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ==================== 优化函数：大量元素（只允许大量肥，固定权重） ====================
def solve_macro_targets(macro_targets, water_for_calc, lib, cf):
    macro_names = [n for n in MACRO_FERTILIZERS if n in lib]

    element_weights = {
        "NO3-N": 100,
        "NH4-N": 100,
        "P": 100,
        "K": 100,
        "Ca": 100,
        "Mg": 180,
        "SO4-S": 20,
        "Urea-N": 20
    }

    prob_macro = pulp.LpProblem("Macro_Opt", pulp.LpMinimize)
    v_macro = {n: pulp.LpVariable(f"macro_{i}", 0, 100) for i, n in enumerate(macro_names)}
    s_macro = {el: pulp.LpVariable(f"sm_{el}", 0) for el in macro_targets.keys()}

    penalty_macro = 0
    for el, target_val in macro_targets.items():
        actual = pulp.lpSum([v_macro[n] * cf * float(lib[n][el]) for n in macro_names])
        net_target = max(0, target_val - water_for_calc.get(el, 0.0))
        prob_macro += actual - net_target <= s_macro[el]
        prob_macro += net_target - actual <= s_macro[el]
        penalty_macro += s_macro[el] * element_weights.get(el, 1)

    prob_macro += pulp.lpSum([v_macro[n] for n in macro_names]) * 0.01 + penalty_macro
    prob_macro.solve(pulp.PULP_CBC_CMD(msg=False))

    status = pulp.LpStatus[prob_macro.status]
    if status in ['Infeasible', 'Undefined', 'Unbounded']:
        return status, {}, element_weights, macro_names

    macro_sol = {n: pulp.value(v_macro[n]) for n in macro_names if pulp.value(v_macro[n]) and pulp.value(v_macro[n]) > 0.001}
    return status, macro_sol, element_weights, macro_names

# ==================== 优化函数：微量元素（只允许微量肥，直接求解） ====================
def solve_micro_targets(micro_targets, lib, cf):
    micro_names = [n for n in MICRO_FERTILIZERS if n in lib]

    prob_micro = pulp.LpProblem("Micro_Opt", pulp.LpMinimize)
    v_micro = {n: pulp.LpVariable(f"micro_{i}", 0, 100) for i, n in enumerate(micro_names)}
    s_micro = {el: pulp.LpVariable(f"sc_{el}", 0) for el in micro_targets.keys()}

    micro_weights = {
        "Fe": 1000,
        "Mn": 1000,
        "Zn": 1000,
        "Cu": 1000,
        "B": 1000,
        "Mo": 1000
    }

    penalty_micro = 0
    for el, target_val in micro_targets.items():
        actual = pulp.lpSum([v_micro[n] * cf * float(lib[n][el]) for n in micro_names])
        prob_micro += actual - target_val <= s_micro[el]
        prob_micro += target_val - actual <= s_micro[el]
        penalty_micro += s_micro[el] * micro_weights.get(el, 1)

    prob_micro += pulp.lpSum([v_micro[n] for n in micro_names]) * 0.001 + penalty_micro
    prob_micro.solve(pulp.PULP_CBC_CMD(msg=False))

    status = pulp.LpStatus[prob_micro.status]
    if status in ['Infeasible', 'Undefined', 'Unbounded']:
        return status, {}, micro_names

    micro_sol = {n: pulp.value(v_micro[n]) for n in micro_names if pulp.value(v_micro[n]) and pulp.value(v_micro[n]) > 0.001}
    return status, micro_sol, micro_names

# ==================== 侧边栏（系统参数 + 原水数据） ====================
with st.sidebar:
    st.header("⚙️ 系统参数")
    tank_vol = st.number_input("母液桶体积(L)", min_value=1.0, value=1000.0, step=100.0)
    dosing_rate = st.number_input("吸肥比例(%)", min_value=0.01, value=0.53, step=0.01) / 100
    ec_calib = st.slider("EC 修正系数", 0.8, 1.4, 1.08, 0.01)

    st.divider()
    st.header("💧 原水数据")
    w_elements = ["NO3-N","NH4-N","P","K","Ca","Mg","SO4-S","Fe","Mn","Zn","Cu","B","Mo"]
    w_data = {el: st.number_input(el, min_value=0.0, value=0.0, step=0.1, key=f"w_{el}") for el in w_elements}
    w_data["HCO3"] = st.number_input("HCO3 (碳酸氢根) ppm", min_value=0.0, value=0.0, step=1.0)
    w_data["EC"] = st.number_input("原水 EC", min_value=0.0, value=0.05, step=0.01)
    w_data["pH"] = st.number_input("原水 pH", min_value=0.0, max_value=14.0, value=7.0, step=0.1)

# ==================== 主界面 Tabs ====================
st.title("🧪 营养液计算系统 v1.7")

tab1, tab_acid, tab2, tab3 = st.tabs([
    "🏗️ 肥料库",
    "💧 调酸设置",
    "🔎 配方回测",
    "🚀 结果回推"
])

# Tab1：肥料库
with tab1:
    st.session_state.fert_lib = st.data_editor(
        st.session_state.fert_lib,
        num_rows="dynamic",
        use_container_width=True
    )

    st.divider()
    c1, c2 = st.columns(2)
    with c1:
        st.subheader("大量肥名单")
        st.write(MACRO_FERTILIZERS)
    with c2:
        st.subheader("微量肥名单")
        st.write(MICRO_FERTILIZERS)

# Tab：调酸设置
with tab_acid:
    st.header("💧 调酸设置")
    acid_mode = st.selectbox("调酸模式", ["不调酸", "调酸"], index=0, key="acid_mode")
    target_pH = st.slider("目标 pH", 4.0, 7.0, 5.5, 0.1, key="target_pH")

    acid_options_map = {
        "磷酸 (H3PO4)": ["75%", "80%", "85%"],
        "硫酸 (H2SO4)": ["50%", "98%"],
        "硝酸 (HNO3)": ["30%", "40%", "55%", "68%"]
    }

    if "acid_list" not in st.session_state:
        st.session_state.acid_list = [
            {"acid_type": "磷酸 (H3PO4)", "conc_label": "85%", "share": 100.0, "enabled": True}
        ]

    if acid_mode == "调酸":
        st.subheader("酸液组合设置")

        if st.button("➕ 添加一种酸"):
            st.session_state.acid_list.append({
                "acid_type": "磷酸 (H3PO4)",
                "conc_label": "85%",
                "share": 0.0,
                "enabled": True
            })

        delete_idx = None
        for i, acid in enumerate(st.session_state.acid_list):
            st.markdown(f"**酸液 {i+1}**")
            c1, c2, c3, c4, c5 = st.columns([2.2, 1.4, 1.2, 1.0, 0.8])

            acid_types = list(acid_options_map.keys())
            current_type = acid.get("acid_type", "磷酸 (H3PO4)")
            if current_type not in acid_types:
                current_type = "磷酸 (H3PO4)"

            acid["acid_type"] = c1.selectbox(
                f"酸种_{i}",
                acid_types,
                index=acid_types.index(current_type),
                key=f"acid_type_{i}"
            )

            valid_concs = acid_options_map[acid["acid_type"]]
            current_conc = acid.get("conc_label", valid_concs[0])
            if current_conc not in valid_concs:
                current_conc = valid_concs[0]

            acid["conc_label"] = c2.selectbox(
                f"浓度_{i}",
                valid_concs,
                index=valid_concs.index(current_conc),
                key=f"acid_conc_{i}"
            )

            acid["share"] = c3.number_input(
                f"比例%_{i}",
                min_value=0.0,
                max_value=100.0,
                value=float(acid.get("share", 0.0)),
                step=1.0,
                key=f"acid_share_{i}"
            )

            acid["enabled"] = c4.checkbox(
                f"启用_{i}",
                value=acid.get("enabled", True),
                key=f"acid_enable_{i}"
            )

            if c5.button("删除", key=f"acid_del_{i}"):
                delete_idx = i

        if delete_idx is not None:
            st.session_state.acid_list.pop(delete_idx)
            st.rerun()

    (
        w_data_calc_preview,
        base_water_preview,
        acid_additions_preview,
        current_hco3_meq,
        needed_meq_total,
        target_residual_meq,
        acid_L_total,
        acid_detail_rows
    ) = get_water_for_calc(w_data, dosing_rate, tank_vol)

    st.info(
        f"原水 pH: {round(w_data.get('pH', 0.0), 2)} | "
        f"原水碳酸氢根碱度: {round(current_hco3_meq, 2)} meq/L"
    )

    raw_ph = w_data.get("pH", 0.0)
    raw_hco3 = w_data.get("HCO3", 0.0)
    if raw_ph > 7.5 and raw_hco3 < 50:
        st.warning("⚠️ 原水 pH 偏高但 HCO3 不高，建议复测碱度，或检查是否因曝气/失CO2导致 pH 偏高。")
    if raw_ph < 6.5 and raw_hco3 > 150:
        st.warning("⚠️ 原水 pH 偏低但 HCO3 偏高，数据组合异常，建议复检原水。")

    if acid_mode == "调酸":
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("总需中和碱度", f"{round(needed_meq_total, 2)} meq/L")
        c2.metric("总单桶加酸量", f"{round(acid_L_total, 3)} L")
        c3.metric("调酸后 HCO3", f"{round(w_data_calc_preview.get('HCO3', 0.0), 1)} ppm")
        c4.metric("目标残余碱度", f"{round(target_residual_meq, 2)} meq/L")

        add_cols = st.columns(3)
        add_cols[0].metric("P 增加", f"{round(acid_additions_preview.get('P', 0.0), 4)} ppm")
        add_cols[1].metric("NO3-N 增加", f"{round(acid_additions_preview.get('NO3-N', 0.0), 4)} ppm")
        add_cols[2].metric("SO4-S 增加", f"{round(acid_additions_preview.get('SO4-S', 0.0), 4)} ppm")

        if acid_detail_rows:
            st.subheader("各酸贡献明细")
            st.dataframe(pd.DataFrame(acid_detail_rows), use_container_width=True, hide_index=True)
        else:
            st.warning("⚠️ 已开启调酸，但没有有效酸液比例，请设置至少一种酸的比例大于 0。")
    else:
        st.success("✅ 不调酸，无需添加酸液")

    st.caption(
        f"✅ 当前设置：模式={acid_mode} | 目标pH={target_pH} | "
        f"残余碱度={round(target_residual_meq,2)} meq/L | "
        f"HCO3={round(w_data_calc_preview.get('HCO3',0.0),1)} ppm"
    )
    st.warning("⚠️ 实际调酸受水温、缓冲体系和有机质影响，建议先做小杯滴定实验验证。")

# Tab2：配方回测
with tab2:
    names = st.session_state.fert_lib.index.tolist()
    inputs = {}
    cols = st.columns(3)
    for i, n in enumerate(names):
        with cols[i % 3]:
            inputs[n] = st.number_input(f"{n}(kg)", min_value=0.0, step=0.1, key=f"t2_{n}")

    st.info("💡 调酸参数已在【💧 调酸设置】页统一设置，点击下方按钮即可使用最新参数")

    if st.button("开始分析"):
        (
            water_for_calc,
            base_water,
            acid_additions,
            _current_hco3_meq,
            _needed_meq_total,
            _target_residual_meq,
            _acid_L_total,
            acid_detail_rows
        ) = get_water_for_calc(w_data, dosing_rate, tank_vol)

        r, tn, m, e, sc, sa = safe_calc(inputs, tank_vol, dosing_rate, water_for_calc, ec_calib)
        show_results(
            r, tn, m, e, sc, sa, inputs,
            base_water=base_water,
            acid_additions=acid_additions,
            acid_rows=acid_detail_rows,
            raw_water=w_data
        )

# Tab3：结果回推
with tab3:
    st.info("💡 当前模式：第一阶段只允许大量肥；第二阶段只允许微量肥直接求解；固定权重，无模式选择。")
    st.caption("固定权重：Mg 优先，SO4-S 适度放松。")

    d1, d2, d3, d4 = st.columns(4)
    tg = {
        "NO3-N": d1.number_input("目标 NO3-N", 0.0, 300.0, 100.0),
        "NH4-N": d1.number_input("目标 NH4-N", 0.0, 300.0, 50.0),
        "P": d2.number_input("目标 P", 0.0, 100.0, 40.0),
        "K": d2.number_input("目标 K", 0.0, 400.0, 180.0),
        "Ca": d3.number_input("目标 Ca", 0.0, 200.0, 80.0),
        "Mg": d3.number_input("目标 Mg", 0.0, 100.0, 30.0),
        "SO4-S": d4.number_input("目标 SO4-S", 0.0, 200.0, 40.0),
        "Fe": d4.number_input("目标 Fe", 0.000, 10.0, 0.0),
        "Mn": d1.number_input("目标 Mn", 0.000, 7.0, 0.0),
        "Zn": d2.number_input("目标 Zn", 0.000, 5.0, 0.0),
        "Cu": d3.number_input("目标 Cu", 0.000, 3.0, 0.0),
        "B": d4.number_input("目标 B", 0.000, 8.0, 0.0),
        "Mo": d2.number_input("目标 Mo", 0.000, 3.0, 0.0),
        "Urea-N": d1.number_input("目标 Urea-N", 0.00, 100.0, 0.0)
    }

    if st.button("🚀 求解最优投料"):
        lib = st.session_state.fert_lib.fillna(0.0).to_dict('index')
        cf = (1_000_000 * dosing_rate) / tank_vol

        (
            water_for_calc,
            base_water,
            acid_additions,
            _current_hco3_meq,
            _needed_meq_total,
            _target_residual_meq,
            _acid_L_total,
            acid_detail_rows
        ) = get_water_for_calc(w_data, dosing_rate, tank_vol)

        macro_targets = {
            "NO3-N": tg["NO3-N"],
            "NH4-N": tg["NH4-N"],
            "P": tg["P"],
            "K": tg["K"],
            "Ca": tg["Ca"],
            "Mg": tg["Mg"],
            "SO4-S": tg["SO4-S"],
            "Urea-N": tg["Urea-N"]
        }

        micro_targets = {
            "Fe": tg["Fe"],
            "Mn": tg["Mn"],
            "Zn": tg["Zn"],
            "Cu": tg["Cu"],
            "B": tg["B"],
            "Mo": tg["Mo"]
        }

        macro_status, macro_sol, macro_weights, macro_names = solve_macro_targets(
            macro_targets=macro_targets,
            water_for_calc=water_for_calc,
            lib=lib,
            cf=cf
        )

        if macro_status in ['Infeasible', 'Undefined', 'Unbounded']:
            st.error("❌ 大量元素无解，请调整目标值、酸方案或肥料库。")
        else:
            micro_status, micro_sol, micro_names = solve_micro_targets(
                micro_targets=micro_targets,
                lib=lib,
                cf=cf
            )

            final_sol = dict(macro_sol)
            for k, v in micro_sol.items():
                final_sol[k] = final_sol.get(k, 0.0) + v

            st.success("✅ 已完成：大量元素与微量元素按肥料类别分阶段求解")

            c1, c2, c3 = st.columns(3)
            c1.metric("大量阶段允许肥料数", len(macro_names))
            c2.metric("微量阶段允许肥料数", len(micro_names))
            c3.metric("最终投料肥料数", len(final_sol))

            st.subheader("第一阶段：大量元素方案（仅大量肥）")
            st.caption(f"允许肥料：{', '.join(macro_names)}")
            if macro_sol:
                macro_df = pd.DataFrame(list(macro_sol.items()), columns=["肥料", "投料 kg"])
                macro_df["投料 kg"] = macro_df["投料 kg"].round(4)
                st.dataframe(macro_df, use_container_width=True, hide_index=True)
            else:
                st.info("大量元素阶段无投料结果。")

            st.subheader("第二阶段：微量元素方案（仅微量肥，直接求解）")
            st.caption(f"允许肥料：{', '.join(micro_names)}")
            if micro_sol:
                micro_df = pd.DataFrame(list(micro_sol.items()), columns=["肥料", "投料 kg"])
                micro_df["投料 kg"] = micro_df["投料 kg"].round(4)
                st.dataframe(micro_df, use_container_width=True, hide_index=True)
            else:
                if micro_status in ['Infeasible', 'Undefined', 'Unbounded']:
                    st.warning("⚠️ 微量元素直接求解无解。")
                else:
                    st.info("本次微量元素无需额外补充。")

            r, tn, m, e, sc, sa = safe_calc(final_sol, tank_vol, dosing_rate, water_for_calc, ec_calib)
            show_results(
                r, tn, m, e, sc, sa, final_sol,
                base_water=base_water,
                acid_additions=acid_additions,
                acid_rows=acid_detail_rows,
                raw_water=w_data
            )

            st.divider()
            st.subheader("📊 大量元素目标 vs 实际对比（总浓度）")
            comparison_data = []
            for el, target in macro_targets.items():
                actual_val = r.get(el, 0.0)
                diff = actual_val - target
                pct_error = (diff / target * 100) if target > 0 else 0.0
                comparison_data.append({
                    "元素": el,
                    "目标 ppm": round(target, 4),
                    "实际 ppm": round(actual_val, 4),
                    "差值": round(diff, 4),
                    "%偏差": f"{round(pct_error, 1)}%"
                })
            comp_df = pd.DataFrame(comparison_data)

            def color_deviation(val):
                try:
                    vv = float(str(val).strip('%'))
                    if abs(vv) <= 2:
                        return 'background-color: #d4edda; color: #155724'
                    elif abs(vv) <= 5:
                        return 'background-color: #fff3cd; color: #856404'
                    else:
                        return 'background-color: #f8d7da; color: #721c24'
                except:
                    return ''

            styled_df = comp_df.style.map(color_deviation, subset=['%偏差']).format({"%偏差": "{}"}).set_properties(**{'text-align': 'center'})
            st.dataframe(styled_df, use_container_width=True, hide_index=True)

            st.subheader("🔬 微量元素目标 vs 实际对比（微量肥直接求解）")
            micro_compare = []
            for el, target in micro_targets.items():
                actual_val = r.get(el, 0.0)
                diff = actual_val - target
                pct_error = (diff / target * 100) if target > 0 else 0.0
                micro_compare.append({
                    "元素": el,
                    "目标 ppm": round(target, 4),
                    "最终实际 ppm": round(actual_val, 4),
                    "差值": round(diff, 4),
                    "%偏差": f"{round(pct_error, 1)}%"
                })
            micro_df = pd.DataFrame(micro_compare)
            styled_micro_df = micro_df.style.map(color_deviation, subset=['%偏差']).format({"%偏差": "{}"}).set_properties(**{'text-align': 'center'})
            st.dataframe(styled_micro_df, use_container_width=True, hide_index=True)

st.caption("百瑞 Blueberry Pro v1.2")
